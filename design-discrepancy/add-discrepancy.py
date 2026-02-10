#!/usr/bin/env python3
"""
Add new row to Design Discrepancy Excel
Usage: ./add-discrepancy.py --title "..." --description "..." --screenshot "..." --figma-url "..." --system-url "..."
"""

import openpyxl
from openpyxl.styles import Font, Alignment
from openpyxl.drawing.image import Image as ExcelImage
import shutil
from datetime import datetime
import argparse
import os

EXCEL_PATH = os.path.expanduser('~/Documents/wiley/OneDrive/OneDrive - Wiley/RPM/Design-Discrepancy.xlsx')

def backup_excel():
    """Create timestamped backup"""
    timestamp = datetime.now().strftime('%Y%m%d-%H%M%S')
    backup_dir = os.path.dirname(EXCEL_PATH)
    backup_path = os.path.join(backup_dir, f'Design-Discrepancy-{timestamp}.xlsx')
    shutil.copy2(EXCEL_PATH, backup_path)
    print(f"✅ Backup: Design-Discrepancy-{timestamp}.xlsx")
    return backup_path

def get_image_dimensions(img_path):
    """Get original image dimensions"""
    import subprocess
    result = subprocess.run(
        ['sips', '-g', 'pixelWidth', '-g', 'pixelHeight', img_path],
        capture_output=True, text=True
    )
    lines = result.stdout.strip().split('\n')
    width = int([l for l in lines if 'pixelWidth' in l][0].split(':')[1].strip())
    height = int([l for l in lines if 'pixelHeight' in l][0].split(':')[1].strip())
    return width, height

def add_row(title, description, screenshot_path, figma_url, system_url, proposed_solution='', ux_considerations=''):
    """Add new discrepancy row"""
    
    # Backup
    backup_excel()
    
    # Load
    wb = openpyxl.load_workbook(EXCEL_PATH)
    ws = wb.active
    
    next_row = ws.max_row + 1
    row_id = next_row - 1
    
    print(f"\n📝 Adding row {next_row} (ID: {row_id})...")
    
    # Data
    ws[f'A{next_row}'] = row_id
    ws[f'B{next_row}'] = title
    ws[f'C{next_row}'] = 'Undefined'
    ws[f'D{next_row}'] = 'New'
    ws[f'E{next_row}'] = 'To triage'
    ws[f'F{next_row}'] = description
    
    # Links (blue + underline)
    ws[f'G{next_row}'] = f'=HYPERLINK("{figma_url}", "Figma")'
    ws[f'G{next_row}'].font = Font(color='00004B', underline='single')
    
    ws[f'H{next_row}'] = f'=HYPERLINK("{system_url}", "System")'
    ws[f'H{next_row}'].font = Font(color='00004B', underline='single')
    
    ws[f'J{next_row}'] = proposed_solution
    ws[f'K{next_row}'] = ux_considerations
    
    # Format: top align + wrap
    for cell in ws[next_row]:
        cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
    
    # Image
    if screenshot_path and os.path.exists(screenshot_path):
        orig_width, orig_height = get_image_dimensions(screenshot_path)
        target_width = 400
        target_height = int(target_width * (orig_height / orig_width))
        
        img = ExcelImage(screenshot_path)
        img.width = target_width
        img.height = target_height
        
        ws.add_image(img, f'I{next_row}')
        
        # Adjust row height
        ws.row_dimensions[next_row].height = target_height * 0.75
        
        print(f"  ✅ Image: {target_width}x{target_height}px")
    
    # Save
    wb.save(EXCEL_PATH)
    print(f"\n✅ Row {next_row} added!")
    print(f"  Title: {title}")
    print(f"  ID: {row_id}")
    
    # Reopen Excel
    os.system('killall "Microsoft Excel" 2>/dev/null')
    os.system(f'sleep 1 && open -a "Microsoft Excel" "{EXCEL_PATH}"')

if __name__ == '__main__':
    parser = argparse.ArgumentParser(description='Add Design Discrepancy row')
    parser.add_argument('--title', required=True, help='Short title')
    parser.add_argument('--description', required=True, help='Detailed description')
    parser.add_argument('--screenshot', required=True, help='Path to screenshot')
    parser.add_argument('--figma-url', required=True, help='Figma link')
    parser.add_argument('--system-url', required=True, help='System link')
    parser.add_argument('--proposed-solution', default='', help='How to fix')
    parser.add_argument('--ux-considerations', default='', help='UX questions/concerns')
    
    args = parser.parse_args()
    
    add_row(
        args.title,
        args.description,
        args.screenshot,
        args.figma_url,
        args.system_url,
        args.proposed_solution,
        args.ux_considerations
    )
